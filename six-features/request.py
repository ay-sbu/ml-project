import requests
import openpyxl

# torob api base url
base_url = 'https://api.torob.com/v4/base-product/details-log-click/?prk='

# read prepared links and make an array
links_file = open('links.txt', 'r')
links_str = links_file.read()
links = links_str.split('\n\n')

# preparing excel sheet to write in file
path = 'api-features.xlsx'
wb = openpyxl.Workbook()
sheet = wb.active

# columns and their numbers
column_number = {  
    'link': 1, 
    'cpu': 2, 
    'ram': 3, 
    'hdd': 4,
    'ssd': 5,
    'graphic_ram': 6,
    'screen_size': 7,
    'stock_status': 8,
    'price': 9
}

# iterate on all links
for i in range(len(links)):
    print()
    print('link ', str(i))

    # write link in first column
    sheet.cell(row=i+1, column=column_number['link']).value = links[i]

    try:
        link_parts = links[i].split('/')
        link_id = link_parts[4]
        
        url = base_url + link_id

        resp = requests.get(url=url)
        data = resp.json() 

        attributes = data['attributes']

        for attribute in attributes:
            if attribute in column_number:
                sheet.cell(row=i+1, column=column_number[attribute]).value = attributes[attribute]

    except Exception as e:
        print()
        print('We Have ERRORR: ')
        print(e)


wb.save(path)

print()
print('success!')



    