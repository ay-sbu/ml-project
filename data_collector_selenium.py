# seleniums
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.chrome.options import Options

# openpyxl
import openpyxl

# built-in
import time
import os

# preparing excel sheet to write in file
path = 'dataset.xlsx'
wb = openpyxl.Workbook()
sheet = wb.active

# write first column title
c1 = sheet.cell(row = 1, column = 1)
c1.value = "link"

# install needed drivers and add it to PATH
os.environ['PATH'] += r"C:/SeleniumDrivers"

# set your setting for driver
chrome_options = Options()
chrome_options.add_argument("--disable-web-security")
chrome_options.add_argument("--allow-running-insecure-content")
chrome_options.add_experimental_option("detach", True)
driver = webdriver.Chrome(options=chrome_options)
driver.implicitly_wait(4)

# read prepared links and make an array
links_file = open('links.txt', 'r')
links_str = links_file.read()
links = links_str.split('\n\n')

# debug
# print("len(links)", len(links))s
# for i in range(10):
# 	print(i)
# 	print(links[i])

founded_titles = []

start = 0
end = 1100

for i in range(start, end):
	# debug
	print()
	print(i, links[i])

	# write link as first property
	sheet_cell = sheet.cell(row=i+2, column=1)
	sheet_cell.value = links[i]

	# price
	try:
		price_element = driver.find_element(By.CSS_SELECTOR, 'div[style="display:inline-block"].jsx-63b317fab2efbae')
		price = price_element.get_attribute("innerHTML")
		sheet_cell = sheet.cell(row=i+2, column=2)
		sheet_cell.value = price
	except:
		print()
		print("price not found!")
		sheet_cell = sheet.cell(row=i+2, column=2)
		sheet_cell.value = 'null'


	# collecting data from link page
	driver.get(links[i])

	titles_element = driver.find_elements(By.CSS_SELECTOR, ".jsx-5b5c456cc255c2dc.detail-title")
	values_element = driver.find_elements(By.CSS_SELECTOR, ".jsx-5b5c456cc255c2dc.detail-value")

	titles = [j.get_attribute("innerHTML") for j in titles_element]
	values = [j.get_attribute("innerHTML") for j in values_element]

	# debug
	# print('--------------')
	# for p in range(len(titles)):
	# 	print(titles[p], ": ", values[p])
	# print('--------------')
	

	for j in range(len(titles)):
		title_index = -1
		for k in range(len(founded_titles)):
			if titles[j] == founded_titles[k]:
				title_index = k
				break

		# title does not exist before
		if title_index == -1:
			title_index = len(founded_titles)

			sheet_cell = sheet.cell(row=1, column=title_index+3)
			sheet_cell.value = titles[j]

			founded_titles.append(titles[j])
		
		# writing value into correct cell
		sheet_cell = sheet.cell(row=i+2, column=title_index+3)
		sheet_cell.value = values[j]

wb.save(path)