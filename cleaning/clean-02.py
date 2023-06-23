# seleniums
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.chrome.options import Options

# openpyxl
import openpyxl

# built-in
import time
import os

# preparing excel sheet to read from file
path = 'clean-02.xlsx'
wb = openpyxl.load_workbook(filename=path)
sheet = wb.active

# install needed drivers and add it to PATH
os.environ['PATH'] += r"C:/SeleniumDrivers"

# set your setting for driver
driver = webdriver.Firefox()
driver.implicitly_wait(4)

price_col = 2
link_col = 1

start = 900
end = 1100

for i in range(start, end):
	price = sheet.cell(row=i, column=price_col).value
	if ('فروشنده' in price) or ('_' in price) or ('null' in price):
		
		# go to link and update price
		driver.get(sheet.cell(row=i, column=link_col).value)

		price_element = driver.find_elements(By.CSS_SELECTOR, ".jsx-63b317fab2efbae.buy_box_text.ellipsis")
		
		try:
			sheet.cell(row=i, column=price_col).value = price_element[1].get_attribute("innerHTML")
		except:
			print('some errors!')

		print(i)

new_path = 'clean-02.xlsx'
wb.save(new_path)

		





